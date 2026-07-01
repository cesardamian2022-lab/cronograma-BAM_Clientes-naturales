import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import io
import copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import tempfile
import subprocess
import sys

# Configuración inicial de la página web
st.set_page_config(page_title="Emisión de Bonos Sostenibles", layout="wide", page_icon="🌳")
st.title("📄 Estructurador de Cronograma de Pagos - BAM")

# --- FUNCIONES DE APOYO (FECHAS EN ESPAÑOL) ---
def fecha_a_espanol(fecha):
    if pd.isnull(fecha) or isinstance(fecha, str):
        return fecha
    dias = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]
    
    dia = fecha.day
    dia_semana = dias[fecha.weekday()]
    mes = meses[fecha.month - 1]
    anio = fecha.year
    return f"{dia_semana}, {dia} de {mes} de {anio}"

# --- 1. PANEL DE ENTRADA ---
with st.sidebar:
    st.header("⚙️ Parámetros de la Inversión")
    
    # Input para que definas la parte complementaria del archivo
    complemento_archivo = st.text_input("Complemento para Archivo", "Nombre_del_Fondo_o_Cliente")

    # Variables de control vacías al inicio de la barra lateral (¡AQUÍ INICIALIZAMOS EMISIÓN!)
    codigo_serie = ""
    moneda_forzada = None
    emision = "" # <-- Corrección clave: Inicializamos la variable por defecto

    programas = [
        "1° Programa Bonos Sostenibles",
        "1° Programa Privado de IRD",
        "1° Emisión Individual de Bonos USD",
        "2° Emisión Individual de Bonos PEN",
        "1° Emisión de Papeles Comerciales (USD-PEN)"
    ]
    
    programa = st.selectbox("Programa", programas)
    
    if programa == "1° Programa Bonos Sostenibles":
        emision = st.selectbox("Emisión", ["5ta Emisión", "6ta Emisión", "7ma Emisión"])
        if emision == "5ta Emisión": codigo_serie = "5E BS"
        elif emision == "6ta Emisión": codigo_serie = "6E BS"
        elif emision == "7ma Emisión": 
            codigo_serie = "7E BS"
            moneda_forzada = "PEN" # Fuerza Soles
            
    elif programa == "1° Programa Privado de IRD":
        emision = st.selectbox("Emisión", ["1° Emisión-Bonos", "2° Emisión-Papeles Comerciales", "3° Emisión-Bonos"])
        if emision == "1° Emisión-Bonos": codigo_serie = "1E IRD"
        elif emision == "2° Emisión-Papeles Comerciales": codigo_serie = "2E IRD"
        elif emision == "3° Emisión-Bonos": codigo_serie = "3E IRD"
        
    elif programa == "1° Emisión Individual de Bonos USD":
        # Corrección: Ahora asignamos el selectbox a la variable 'emision'
        emision = st.selectbox("Emisión", ["No aplica (Emisión Individual)"], disabled=True)
        codigo_serie = "1E IND"
        moneda_forzada = "USD" # Fuerza Dólares
        
    elif programa == "2° Emisión Individual de Bonos PEN":
        # Corrección: Ahora asignamos el selectbox a la variable 'emision'
        emision = st.selectbox("Emisión", ["No aplica (Emisión Individual)"], disabled=True)
        codigo_serie = "2E IND"
        moneda_forzada = "PEN" # Fuerza Soles
        
    elif programa == "1° Emisión de Papeles Comerciales (USD-PEN)":
        # Corrección: Ahora asignamos el selectbox a la variable 'emision'
        emision = st.selectbox("Emisión", ["No aplica (Emisión Individual)"], disabled=True)
        codigo_serie = "1PC"
        # Queda libre para que el usuario elija, pero con la alerta de riesgo operativo
    
    inversionista = st.text_input("I7: Nombre Inversionista", "")
    Letra_serie = st.text_input("I7: Letra de la Serie", "")

    # NUEVOS INPUTS DE DOCUMENTO
    tipo_documento = st.selectbox("Tipo de documento", ["DNI", "RUC", "CE","CI", "Pasaporte"], index=1)
    numero_documento = st.text_input("Número de documento", "")
    
    tipo_plazo = st.radio("Definición de Plazo", ["Días Exactos", "Años Exactos"])
    if tipo_plazo == "Días Exactos":
        plazo_input = st.number_input("I8: Plazo en Días", min_value=30, value=731)
    else:
        plazo_input = st.number_input("Cantidad de Años", min_value=1, value=2)
        
    if moneda_forzada:
        # Si las reglas de arriba dictaron una moneda, bloqueamos la caja
        moneda = st.selectbox("I9: Moneda", [moneda_forzada], disabled=True)
        st.info(f"📌 Moneda fijada a {moneda_forzada} por las reglas del programa.")
    else:
        # Si no hay restricción, dejamos elegir libremente
        moneda = st.selectbox("I9: Moneda", ["USD", "PEN"])
        
        # Alerta especial solo para Papeles Comerciales
        if programa == "1° Emisión de Papeles Comerciales (USD-PEN)":
            st.warning("⚠️ ATENCIÓN: Verifique que la moneda sea la correcta para esta emisión de Papeles Comerciales.")
    
    monto = st.number_input("I10: Monto a invertir", min_value=1000.0, value=130000.0, step=1000.0)
    
    tipo_tasa = st.selectbox("I11: Tipo de Tasa", ["Anual", "Mensual", "A Vencimiento"])
    tasa_input = st.number_input("Tasa Nominal (%)", value=10.25, step=0.01, format="%.2f") / 100
    
    escoger_pdf = st.selectbox("Generar en Formato PDF", ["NO", "SÍ"], index=0)

    fecha_emision = st.date_input("I12: Fecha de Inversión (Emisión)", datetime(2026, 5, 11))
    frecuencia = st.selectbox("I13: Frecuencia del cupón", ["Trimestral", "Mensual"])
    
    tipo_entidad = st.selectbox("I14: Tipo de Inversionista (Para IR)", [
        "Persona Jurídica Domiciliada (RUC)", 
        "Persona Jurídica No Domiciliada",
        "Persona Natural Domiciliada (DNI)",
        "Persona Natural No Domiciliada"
    ], index=2)

# --- 2. MOTOR DE REGLAS DE NEGOCIO ---
if tipo_plazo == "Años Exactos":
    fecha_redencion = fecha_emision + relativedelta(years=plazo_input)
    plazo_total_dias = (fecha_redencion - fecha_emision).days
else:
    plazo_total_dias = int(plazo_input)
    fecha_redencion = fecha_emision + timedelta(days=plazo_total_dias)

tasa_ir = 0.0
if tipo_entidad in ["Persona Jurídica No Domiciliada", "Persona Natural No Domiciliada"]:
    tasa_ir = 0.0499
elif tipo_entidad == "Persona Natural Domiciliada (DNI)":
    tasa_ir = 0.05

tna = tasa_input if tipo_tasa != "Mensual" else tasa_input * 12

def calcular_primer_25(fecha_emision, frecuencia):
    dias_teoricos = 90 if frecuencia == "Trimestral" else 30
    fecha_objetivo = fecha_emision + timedelta(days=dias_teoricos)
    
    opciones = [
        (fecha_objetivo + relativedelta(months=-1)).replace(day=25),
        fecha_objetivo.replace(day=25),
        (fecha_objetivo + relativedelta(months=1)).replace(day=25)
    ]
    return min(opciones, key=lambda x: abs((x - fecha_objetivo).days))

# --- 3. CONSTRUCCIÓN DEL CRONOGRAMA ---
cronograma = []
fecha_anterior = fecha_emision
meses_salto = 3 if frecuencia == "Trimestral" else 1

fecha_siguiente = calcular_primer_25(fecha_emision, frecuencia)
pago_num = 1

while fecha_siguiente < fecha_redencion:
    dias_periodo = (fecha_siguiente - fecha_anterior).days
    
    if tipo_tasa == "A Vencimiento":
        cupon_bruto = monto * tasa_input
    else:
        cupon_bruto = monto * (tna / 360) * dias_periodo
        
    retencion = cupon_bruto * tasa_ir
    neto = cupon_bruto - retencion
    
    cronograma.append({
        "Pago": f"Cupón {pago_num}",
        "Fecha de vencimiento": fecha_siguiente,
        "Plazo": dias_periodo,
        "Moneda": moneda,
        "Cupón": cupon_bruto,
        "Retención IR": -retencion,
        "Neto a pagar": neto
    })
    
    pago_num += 1
    fecha_anterior = fecha_siguiente
    fecha_siguiente = fecha_siguiente + relativedelta(months=meses_salto)

dias_ultimo_periodo = (fecha_redencion - fecha_anterior).days
if dias_ultimo_periodo > 0:
    cupon_bruto = monto * tasa_input if tipo_tasa == "A Vencimiento" else monto * (tna / 360) * dias_ultimo_periodo
    retencion = cupon_bruto * tasa_ir
    neto = cupon_bruto - retencion
    
    cronograma.append({
        "Pago": f"Cupón {pago_num}",
        "Fecha de vencimiento": fecha_redencion,
        "Plazo": dias_ultimo_periodo,
        "Moneda": moneda,
        "Cupón": cupon_bruto,
        "Retención IR": -retencion,
        "Neto a pagar": neto
    })

# Fila Final del Capital
cronograma.append({
    "Pago": "Capital",
    "Fecha de vencimiento": fecha_redencion,
    "Plazo": None,
    "Moneda": moneda,
    "Cupón": None,
    "Retención IR": None,
    "Neto a pagar": monto
})

df = pd.DataFrame(cronograma)

# --- 4. RENDERIZADO VISUAL EN LA WEB ---
st.write(f"### Cotización a Inversionista: {inversionista}")
col1, col2, col3 = st.columns(3)

# Función de apoyo rápida para dar formato a la métrica pequeña
def metrica_pequena(titulo, valor):
    return f"""
    <div style='line-height: 1.3; margin-bottom: 15px;'>
        <span style='font-size: 12px; color: #666;'>{titulo}</span><br>
        <span style='font-size: 16px; font-weight: 600;'>{valor}</span>
    </div>
    """

# Insertamos el texto usando markdown y permitiendo HTML
col1.markdown(metrica_pequena("Fecha de Emisión", fecha_a_espanol(fecha_emision)), unsafe_allow_html=True)
col2.markdown(metrica_pequena("Fecha de Redención", fecha_a_espanol(fecha_redencion)), unsafe_allow_html=True)
col3.markdown(metrica_pequena("Plazo Exacto", f"{plazo_total_dias} días"), unsafe_allow_html=True)  
# Copiamos para no dañar los números del Excel que generaremos luego
df_mostrar = df.copy()

df_mostrar["Fecha de vencimiento"] = df_mostrar["Fecha de vencimiento"].apply(fecha_a_espanol)
df_mostrar["Plazo"] = df_mostrar["Plazo"].apply(lambda x: f"{int(x)}" if pd.notnull(x) else "")

# Formatear el dinero a 2 decimales para la pantalla
for col in ["Cupón", "Retención IR", "Neto a pagar"]:
    df_mostrar[col] = df_mostrar[col].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")

st.dataframe(df_mostrar, use_container_width=True, hide_index=True)

# Total Neto incluye todos los cupones netos + el capital devuelto
total_neto = df["Neto a pagar"].sum()
st.success(f"**MONTO A RECIBIR: {moneda} {total_neto:,.2f}**")

# --- 5. EXPORTACIÓN A EXCEL (PLANTILLA CORPORATIVA ENMARCADA) ---
def generar_excel_bam(df, inv, plazo_d, moneda, monto, tna, f_emi, f_red, frec, ir, titulo_crono, tipo_doc, num_doc):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    ws.sheet_view.showGridLines = False
    
    # 🎨 Paleta de colores y estilos corporativos
    ft_bold = Font(bold=True)
    ft_white_bold = Font(bold=True, color="FFFFFF") 
    ft_purple_bold = Font(bold=True, color="51154A") 
    fill_dark_green = PatternFill(start_color="196B24", end_color="196B24", fill_type="solid") 
    ft_title = Font(bold=True, size=14) 

    # Títulos base dinámicos
    ws["F3"] = "BOSQUES AMAZÓNICOS S.A."
    ws["F4"] = titulo_crono # <--- Título dinámico inyectado aquí
    ws["F3"].font = ft_title 
    ws["F4"].font = ft_title 
    ws["F3"].alignment = Alignment(horizontal="center")
    ws["F4"].alignment = Alignment(horizontal="center")

    # 🖼️ LOGO BAM
    try:
        logo = OpenpyxlImage('logo_bam.png') 
        logo.width = 195
        logo.height = 55
        ws.add_image(logo, 'I2')
    except FileNotFoundError:
        st.warning("⚠️ No se encontró el archivo de imagen 'logo_bam.png'.")

    # CARACTERÍSTICAS (Ahora incluye Documentos)
    ws["C6"] = "CARACTERISTICAS DE LA INVERSION"
    etiquetas = ["Nombre Inversionista", "Plazo", "Moneda", "Monto a invertir", 
                 "Tasa Nominal Anual (TNA)", "Fecha de Inversión", "Frecuencia del cupón", 
                 "Tasa de Retención de Impuesto", "Fecha de Redención", "Tipo de documento", "Número de documento"]
    valores = [inv, plazo_d, moneda, monto, tna, fecha_a_espanol(f_emi), frec, ir, fecha_a_espanol(f_red), tipo_doc, num_doc]
    
    for i, (et, val) in enumerate(zip(etiquetas, valores)):
        ws.cell(row=7+i, column=3, value=et)
        ws.cell(row=7+i, column=9, value=val)
        
    # FILAS DESPLAZADAS +2
    ws["C19"] = "COTIZACION A INVERSIONISTA"
    ws["C21"] = "MONTO A INVERTIR"
    ws["F21"] = moneda 
    ws["F21"].alignment = Alignment(horizontal="center")
    ws["I21"] = monto
    ws["I21"].number_format = '#,##0.00'
    
    # Formatos de tabla superior
    ws["I10"].number_format = '#,##0.00'
    ws["I11"].number_format = '0.00##%'
    ws["I14"].number_format = '0.00%'       
    
    # Llenado de Cabeceras (Ahora en fila 23)
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=23, column=col_num+2)
        cell.value = header
        cell.alignment = Alignment(horizontal="center")
    borde_limpio = Side(border_style="thin", color="000000")
    borde_sencillo = Border(bottom=borde_limpio)        
    for c in range(3, 10): 
        ws.cell(row=23, column=c).border = borde_sencillo

    # Llenado dinámico de Filas de datos (Inicia en fila 24)
    for r_idx, row_data in enumerate(df.values, 24):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx+2)

            if c_idx == 1:
                cell.value = value
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 2:
                cell.value = fecha_a_espanol(value)
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 3:
                if pd.notnull(value):
                    cell.value = int(value)
                    cell.number_format = '0'
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in [5, 6, 7]:
                if pd.notnull(value):
                    cell.value = round(float(value), 2)
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal="center")

    # Fila final de sumatoria
    fila_actual = 24 + len(df)
    for c_idx in range(1, 8):
        cell = ws.cell(row=fila_actual, column=c_idx+2)
        if c_idx == 1:
            cell.value = "MONTO A RECIBIR"
        elif c_idx == 4:
            cell.value = moneda
            cell.alignment = Alignment(horizontal="center")
        elif c_idx == 7:
            cell.value = f"=SUM(I24:I{fila_actual-1})"
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    # ==========================================
    # 🖼️ MOTOR DE PINTADO Y MARCOS 
    # ==========================================
    for c in range(2, 11):
        cell = ws.cell(row=6, column=c)
        cell.fill = fill_dark_green
        if cell.value: cell.font = ft_white_bold
        
    for r in range(6, fila_actual + 1):
        ws.cell(row=r, column=2).fill = fill_dark_green  
        ws.cell(row=r, column=10).fill = fill_dark_green 

    # Borde derecho de las características (Hasta fila 17)
    for r in range(7, 18):
        cell = ws.cell(row=r, column=9)  
        cell.alignment = Alignment(horizontal="right")

    # Franjas Intermedias desplazadas: Filas 19, 21 y la fila_actual
    for r in [19, 21, fila_actual]:
        for c in range(3, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_dark_green
            if cell.value: cell.font = ft_white_bold
            
            # Cabeceras de tabla están en la fila 23
            cell = ws.cell(row=23, column=c)
            cell.font = ft_purple_bold
            if c in [3, 4]: 
                cell.alignment = Alignment(horizontal="left")
            elif c==9:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    borde_punteado = Side(border_style="hair", color="000000")
    borde_completo = Border(top=borde_punteado, bottom=borde_punteado)
    
    # Filas punteadas de características (De 7 a 17)
    for r in range(7, 18):     
        for c in range(3, 10): 
            ws.cell(row=r, column=c).border = borde_completo
    
    # 📏 AJUSTE DE COLUMNAS EXPANDIDAS
    columnas_ancho = {
        'A': 10.82, 'B': 2, 'C': 11.18, 'D': 27, 'E': 10, 
        'F': 10.5, 'G': 10.9, 'H': 15.2, 'I': 26, 'J': 2
    }
    for col_letter, width in columnas_ancho.items():
        ws.column_dimensions[col_letter].width = width
    
    fuente_global = "Aptos Narrow"
    for row in ws.iter_rows(): 
        for cell in row:
            if cell.font: 
                nueva_fuente = copy.copy(cell.font)
                nueva_fuente.name = fuente_global
                cell.font = nueva_fuente 

    # === AJUSTES DE IMPRESIÓN PARA PDF PERFECTO ===
    ws.print_options.horizontalCentered = True
    ws.print_area = f'A1:K{fila_actual + 2}' 
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True 
    ws.page_setup.fitToWidth = 1   
    ws.page_setup.fitToHeight = 1  
    
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 6. EXPORTACIÓN A PDF (VIA EXCEL NATIVO) ---
def generar_pdf_desde_excel(excel_bytes):
    temp_dir = tempfile.gettempdir()
    # Generamos nombres de archivo únicos para evitar cruces si varios usuarios lo usan a la vez
    temp_excel_path = os.path.join(temp_dir, "temp_crono_nube.xlsx")
    temp_pdf_path = os.path.join(temp_dir, "temp_crono_nube.pdf")
    
    # 1. Guardar el archivo de Excel crudo
    with open(temp_excel_path, "wb") as f:
        f.write(excel_bytes.getvalue())
        
    try:
        # Detectar automáticamente si estamos en Windows (Local) o Linux (Nube)
        if sys.platform == "win32":
            # Ruta de LibreOffice en tu laptop
            ejecutable = r"C:\Program Files\LibreOffice\program\soffice.exe"
        else:
            # Comando estándar en la nube (Streamlit Cloud)
            ejecutable = "libreoffice"

        # Comando unificado
        comando = [
            ejecutable, '--headless', '--convert-to', 'pdf',
            '--outdir', temp_dir, temp_excel_path
        ]
        
        # Ejecutamos el comando y esperamos a que termine
        subprocess.run(comando, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
    except Exception as e:
        st.error(f"Error al generar PDF en la nube: {e}")
        return None
            
    # 3. Leer el PDF generado
    pdf_output = io.BytesIO()
    if os.path.exists(temp_pdf_path):
        with open(temp_pdf_path, "rb") as f:
            pdf_output.write(f.read())
        pdf_output.seek(0)
        
    # 4. Limpiar los archivos temporales del servidor
    try:
        if os.path.exists(temp_excel_path): os.remove(temp_excel_path)
        if os.path.exists(temp_pdf_path): os.remove(temp_pdf_path)
    except:
        pass 
        
    return pdf_output

# --- 7. CONTROL DE DESCARGAS ---
st.write("---")
st.header("💾 Descargar Archivos")

nombre_archivo_limpio = complemento_archivo.replace(' ', '_').replace('/', '-')
nombre_archivo_excel = f"CRONO-{codigo_serie}-{Letra_serie}-{nombre_archivo_limpio}.xlsx"
nombre_archivo_pdf = f"CRONO-{codigo_serie}-{Letra_serie}-{nombre_archivo_limpio}.pdf"

# === NUEVA LÓGICA DE TÍTULO DINÁMICO ===
if emision and "No aplica" not in emision:
    titulo_cronograma = f"CRONOGRAMA: {programa} - {emision}"
else:
    titulo_cronograma = f"CRONOGRAMA: {programa}"

# 1. Generamos el Excel base (siempre se necesita, ya sea para descargar o como molde del PDF)
excel_file = generar_excel_bam(df, inversionista, plazo_total_dias, moneda, monto, tna, fecha_emision, fecha_redencion, frecuencia, tasa_ir, 
    titulo_cronograma, tipo_documento, numero_documento)

col_btn1, col_btn2 = st.columns(2)

with col_btn1:
    st.download_button(
        label="📥 Descargar Cronograma (Excel)",
        data=excel_file,
        file_name=nombre_archivo_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True 
    )

with col_btn2:
    if escoger_pdf == "SÍ":
        # NUEVA LÓGICA: Le pasamos el Excel ya diseñado al generador de PDF
        with st.spinner('Ensamblando PDF corporativo...'):
            pdf_file = generar_pdf_desde_excel(excel_file)
            
        if pdf_file:
            st.download_button(
                label="📥 Descargar Cronograma (PDF)",
                data=pdf_file,
                file_name=nombre_archivo_pdf,
                mime="application/pdf",
                disabled=False,
                use_container_width=True
            )
        else:
            st.error("No se pudo compilar el PDF localmente.")
    else:
        st.download_button(
            label="🚫 Generación de PDF Desactivada",
            data=b"", 
            file_name="vacio.pdf",
            disabled=True,
            help="Selecciona 'SÍ' en la barra lateral (Creación de PDF) para habilitar esta descarga.",
            use_container_width=True
        )
